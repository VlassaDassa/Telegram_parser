from pathlib import Path

from parser.parser import get_ids, get_data

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl import load_workbook





alphabet = ['A', 'B', 'C', 'D', 'E']


def autosize_column(ws, values):
    column = 0
    for row in values:
        ws.column_dimensions[alphabet[column]].width = len(str(row))+5
        column += 1


def autosize_full(ws, values):
    name = []
    salePrice = []
    basePrice = []
    image = []
    category = []

    for i in values:
        row = list(i)
        if row[0] is None:
            row[0] = 'Имя'
        name.append(row[0])
        salePrice.append(str(row[1]))
        basePrice.append(str(row[2]))
        image.append(row[3])
        category.append(row[4])

    width_name = len(max(name, key=len))+5
    width_salePrice = len(max(salePrice, key=len))+5
    width_basePrice = len(max(basePrice, key=len))+5
    width_image = len(max(image, key=len))+5
    width_category = len(max(category, key=len))+5

    ws.column_dimensions['A'].width = width_name
    ws.column_dimensions['B'].width = width_salePrice
    ws.column_dimensions['C'].width = width_basePrice
    ws.column_dimensions['D'].width = width_image
    ws.column_dimensions['E'].width = width_category



def create_excel_book(tgid):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Товары'

    headers = ('Название', 'Текущая цена', 'Обычная цена', 'Картинка', 'Категория')
    ws.append(headers)
    autosize_column(ws, headers)
    ws.row_dimensions[1].height = 25
    for letter in alphabet:
        ws[f'{letter}1'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'{letter}1'].fill = PatternFill(start_color="cccccc", end_color="cccccc", fill_type="solid")
        ws[f'{letter}1'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wb.save(str(Path(str(Path.cwd()), 'data', f'{tgid}.xlsx')))




def insert_to_excel(page, category, tgid):
    wb = load_workbook(str(Path(str(Path.cwd()), 'data', f'{tgid}.xlsx')))
    ws = wb.active

    for row in get_data(get_ids(page, category), page, category, tgid):
        ws.append(row)

    rows = []
    for i in ws.values:
        rows.append(i)
    next_cord = len(rows)
    start_cord = 2

    for i in range(start_cord, next_cord+1):
        for ii in ['B', 'C', 'E']:
            ws[f'{ii}{i}'].alignment = Alignment(horizontal='center', vertical='center')
        for letter in alphabet:
            ws[f'{letter}{i}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


    autosize_full(ws, ws.values)
    wb.save(str(Path(str(Path.cwd()), 'data', f'{tgid}.xlsx')))











